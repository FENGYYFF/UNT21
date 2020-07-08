import unittest
from testdata.testcase_register import test_register
import openpyxl
from openpyxl import load_workbook

# 创建一个suite对象
suite_register = unittest.TestSuite()

# 创建一个xl的workbook对象
workbook = load_workbook('../testdata/data_xl.xlsx')
# 创建一个页签对象
sheet = workbook['login_sheet']  # type: openpyxl.workbook.workbook.Worksheet
# 创建一个列表接受字典数据
testdata = []

# 获取所有的表格对象存储在元祖中
rows = list(sheet.rows)
coloums = list(sheet.columns)

# 列表推导式
title = [cell.value for cell in rows[0]]

for row in rows[1:]:
    content = [cell.value for cell in row]
    zip_case = dict(zip(title,content))
    testdata.append(zip_case)
"""content = []
for row in rows[1:]:
    for cell in row:
        content.append(cell.value)
        zip_case = dict(zip(title,content))
    testdata.append(zip_case)
"""


"""
方式一：
for row in range(1, sheet.max_row + 1):
    # 创建一个字典存储数据
    dict_data = {}
    if row == 1:
        # 存储表头title
        title = []
        for coloum in range(1, sheet.max_column + 1):
            cell_data = sheet.cell(row, coloum).value
            title.append(cell_data)
    else:
        # 存储内容
        content = []
        for coloum in range(1, sheet.max_column + 1):
            cell_data = sheet.cell(row, coloum).value
            content.append(cell_data)
        # 将二者打包合并
        zip_case = zip(title,content)
        # 往testdata列表中新增一条字典数据
        testdata.append(dict(zip_case))

"""
#
#
for i in testdata:
    suite_register.addTest(test_register(eval(i["data"]),eval(i["expect"])))


# print(eval(i["data"]),type(eval(i["expect"])))


# 当前suite已完成用例装填， 准备进入runner
